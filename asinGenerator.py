import openpyxl

def american():
    wb = openpyxl.load_workbook('/Users/main/desktop/code/scraping/dare_stuff/dare_amazon_scraper/reports_pull/us/Sales Diagnostic_Detail View_US (1).xlsx')
    ws = wb['US_Detail View_Sales Diagnostic']
    mylist = []
    for col_cells in ws.iter_cols(min_col=1, max_col=1):
        for cell in col_cells:    
            mylist.append(cell.value)

    US_ASINS = []
    condition = 'B0'
    for i in mylist:
        if condition in i:
            US_ASINS.append(i)
    
    return US_ASINS

def canadian():
    wb = openpyxl.load_workbook('/Users/main/desktop/code/scraping/dare_stuff/dare_amazon_scraper/reports_pull/can/Sales Diagnostic_Detail View_CA.xlsx')
    ws = wb['CA_Detail View_Sales Diagnostic']
    mylist = []
    for col_cells in ws.iter_cols(min_col=1, max_col=1):
        for cell in col_cells:    
            mylist.append(cell.value)

    CA_ASINS = []
    condition = 'B0'
    for i in mylist:
        if condition in i:
            CA_ASINS.append(i)
    
    return CA_ASINS
  
    



